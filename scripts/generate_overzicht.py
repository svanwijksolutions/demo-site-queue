#!/usr/bin/env python3
"""Genereert overzicht.xlsx uit companies.json. Draait automatisch als onderdeel
van de nachtelijke demo-site pipeline, telkens als companies.json wijzigt.

Naast het hoofdoverzicht (tab "Overzicht") maakt dit script een tweede tab
"Updategeschiedenis": een chronologische lijst (nieuwste eerst) van elke keer dat
een site is gebouwd of aangepast, met de datum en een korte omschrijving van wat er
is veranderd. Zo kan Sem in één oogopslag zien wanneer er iets is bijgewerkt (bijv.
na feedback) en waarop ze kan controleren. De brondata staat per bedrijf in het
`updates`-veld in companies.json (een lijst met {datum, wat, commit?, live_url?})."""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
COMPANIES_JSON = ROOT / "companies.json"
OUTPUT_XLSX = ROOT / "overzicht.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILLS = {
    "done": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "in_progress": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "needs_review": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "pending": PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"),
}

COLUMNS = [
    ("Bedrijfsnaam", 26),
    ("Contactpersoon", 22),
    ("Branche", 30),
    ("Regio", 24),
    ("Status", 14),
    ("Repo", 22),
    ("Live URL", 42),
    ("Pitchmail klaar", 14),
    ("Pitchmail verzonden", 16),
    ("Verzonden op", 14),
    ("Klantreactie", 18),
    ("Follow-up op", 14),
    ("Telefoon", 16),
    ("E-mail", 26),
    ("Toegevoegd op", 14),
    ("Afgerond op", 14),
    ("Laatste update", 14),
    ("Laatste wijziging", 50),
]

UPDATE_COLUMNS = [
    ("Bedrijfsnaam", 26),
    ("Datum", 14),
    ("Wat is er veranderd", 78),
    ("Commit", 14),
    ("Live URL", 42),
]

SENT_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
NOT_SENT_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

REACTIE_FILLS = {
    "klant geworden": PatternFill(start_color="A7F3D0", end_color="A7F3D0", fill_type="solid"),
    "interesse": PatternFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid"),
    "afgewezen": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
    "nog geen reactie": PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"),
}


def load_companies():
    data = json.loads(COMPANIES_JSON.read_text(encoding="utf-8"))
    return data["companies"]


def sorted_updates(company):
    """Updates van een bedrijf, oplopend op datum (oudste eerst)."""
    updates = company.get("updates") or []
    return sorted(updates, key=lambda u: u.get("datum") or "")


def latest_update(company):
    ups = sorted_updates(company)
    return ups[-1] if ups else None


def style_header(ws, columns):
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def build_overview_sheet(ws, companies):
    style_header(ws, COLUMNS)

    for row_idx, c in enumerate(companies, start=2):
        contact = c.get("contact", {})
        last = latest_update(c)
        last_datum = last.get("datum", "") if last else (c.get("afgerond_op") or "")
        last_wat = last.get("wat", "") if last else ""
        values = [
            c.get("bedrijfsnaam", ""),
            c.get("contactpersoon", ""),
            c.get("branche", ""),
            c.get("regio", ""),
            c.get("status", ""),
            c.get("repo", ""),
            c.get("live_url") or "",
            "Ja" if c.get("pitch_email_klaar") else "Nee",
            "Ja" if c.get("pitch_verzonden") else "Nee",
            c.get("pitch_verzonden_op") or "",
            c.get("klant_reactie") or "nog geen reactie",
            c.get("follow_up_op") or "",
            contact.get("telefoon", ""),
            contact.get("email", ""),
            c.get("toegevoegd_op", ""),
            c.get("afgerond_op") or "",
            last_datum,
            last_wat,
        ]
        status_fill = STATUS_FILLS.get(c.get("status"), None)
        sent_col_idx = COLUMNS.index(("Pitchmail verzonden", 16)) + 1
        reactie_col_idx = COLUMNS.index(("Klantreactie", 18)) + 1
        wijziging_col_idx = COLUMNS.index(("Laatste wijziging", 50)) + 1
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            wrap = col_idx == wijziging_col_idx
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
            if col_idx == 5 and status_fill:
                cell.fill = status_fill
            if col_idx == sent_col_idx and c.get("pitch_email_klaar"):
                cell.fill = SENT_FILL if c.get("pitch_verzonden") else NOT_SENT_FILL
            if col_idx == reactie_col_idx:
                cell.fill = REACTIE_FILLS.get(c.get("klant_reactie") or "nog geen reactie", REACTIE_FILLS["nog geen reactie"])

    last_row = max(len(companies) + 1, 2)
    last_col_letter = get_column_letter(len(COLUMNS))
    if len(companies) > 0:
        table = Table(displayName="BedrijvenOverzicht", ref=f"A1:{last_col_letter}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
        )
        ws.add_table(table)


def build_updates_sheet(ws, companies):
    style_header(ws, UPDATE_COLUMNS)

    # Verzamel alle update-events over alle bedrijven, nieuwste eerst.
    events = []
    for c in companies:
        for u in sorted_updates(c):
            events.append((c, u))
    events.sort(key=lambda e: e[1].get("datum") or "", reverse=True)

    for row_idx, (c, u) in enumerate(events, start=2):
        values = [
            c.get("bedrijfsnaam", ""),
            u.get("datum", ""),
            u.get("wat", ""),
            u.get("commit", "") or "",
            u.get("live_url") or c.get("live_url") or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            wrap = col_idx == 3
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)

    last_row = max(len(events) + 1, 2)
    last_col_letter = get_column_letter(len(UPDATE_COLUMNS))
    if len(events) > 0:
        table = Table(displayName="Updategeschiedenis", ref=f"A1:{last_col_letter}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
        )
        ws.add_table(table)


def build_workbook(companies):
    wb = Workbook()
    ws = wb.active
    ws.title = "Overzicht"
    build_overview_sheet(ws, companies)

    ws_updates = wb.create_sheet("Updategeschiedenis")
    build_updates_sheet(ws_updates, companies)

    wb.save(OUTPUT_XLSX)
    return OUTPUT_XLSX


def main():
    companies = load_companies()
    path = build_workbook(companies)
    print(f"overzicht.xlsx geschreven met {len(companies)} bedrijven -> {path}")


if __name__ == "__main__":
    sys.exit(main())
